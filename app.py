import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill

st.title("BRI Transaction Database Generator")

rk_file = st.file_uploader("Upload Rekening Koran", type=["xlsx","xls","csv"])
existing_db_file = st.file_uploader("Upload Existing Database (optional)", type=["xlsx"])


# =====================================
# BRI UNIQUE CODE EXTRACTOR
# =====================================
def ambil_kode_unik(text):

    if pd.isna(text):
        return "N/A"

    text = str(text)

    m = re.search(r'BFVA11167000(\d{5})', text)
    if m:
        return m.group(1)

    m = re.search(r'BRIVA11167000(\d{5})', text)
    if m:
        return m.group(1)

    m = re.search(r'NBMB\s(.*?)\sTO', text)
    if m:
        return m.group(1).strip()

    m = re.search(r'301([A-Z\s]+?):', text)
    if m:
        return m.group(1).strip()

    for i in range(10):
        m = re.search(fr'ATM{i} ATM{i} (.*?)  TO', text)
        if m:
            return m.group(1).strip()

    m = re.search(r'FROM (.*?) LA', text)
    if m:
        return m.group(1).strip()

    m = re.search(r'FROM (.*?) ATM', text)
    if m:
        return m.group(1).strip()

    return "N/A"


# =====================================
# DETECT RK SHEET
# =====================================
def detect_rk_sheet(excel):

    for sheet in excel.sheet_names[:10]:

        df = pd.read_excel(excel, sheet_name=sheet, nrows=5)

        cols = [str(c).lower() for c in df.columns]

        if any(x in c for c in cols for x in ["uraian","description","deskripsi","transaksi"]):
            return sheet

    return excel.sheet_names[0]


# =====================================
# DETECT HEADER ROW
# =====================================
def detect_header(excel, sheet):

    preview = pd.read_excel(excel, sheet_name=sheet, header=None, nrows=20)

    for i in range(20):

        row = preview.iloc[i].astype(str).str.lower()

        if "id" in row.values:
            return i

    return 0


# =====================================
# DETECT ID COLUMN
# =====================================
def detect_id_column(df):

    for col in df.columns:
        if str(col).strip().lower() == "id":
            return col

    for col in df.columns:
        if "id" in str(col).strip().lower():
            return col

    raise Exception("Kolom ID tidak ditemukan")


# =====================================
# DETECT TRANSACTION COLUMN
# =====================================
def detect_desc_column(df):

    for col in df.columns:

        name = str(col).lower()

        if any(x in name for x in ["uraian","description","deskripsi","transaksi"]):
            return col

    # fallback: kolom teks terpanjang
    lengths = df.astype(str).apply(lambda x: x.str.len().mean())

    return lengths.idxmax()


# =====================================
# GROUP CONFLICT
# =====================================
def group_conflict(df):

    normal_rows = []
    conflict_rows = []

    if df.empty:

        return (
            pd.DataFrame(columns=["ID","KODE_UNIK","Uraian Transaksi"]),
            pd.DataFrame(columns=["ID","KODE_UNIK","Uraian Transaksi"])
        )

    grouped = df.groupby("KODE_UNIK")

    for kode, group in grouped:

        ids = sorted(group["ID"].dropna().unique())
        uraian = group["Uraian Transaksi"].dropna().tolist()

        if len(ids) <= 1:

            normal_rows.append({
                "ID": ids[0] if ids else None,
                "KODE_UNIK": kode,
                "Uraian Transaksi": uraian[0] if uraian else None
            })

        else:

            conflict_rows.append({
                "ID": " ; ".join(map(str,ids)),
                "KODE_UNIK": kode,
                "Uraian Transaksi": " ; ".join(uraian)
            })

    return (
        pd.DataFrame(normal_rows, columns=["ID","KODE_UNIK","Uraian Transaksi"]),
        pd.DataFrame(conflict_rows, columns=["ID","KODE_UNIK","Uraian Transaksi"])
    )


# =====================================
# MAIN PROCESS
# =====================================
if rk_file:

    try:

        if rk_file.name.endswith(".csv"):

            df = pd.read_csv(rk_file, sep=None, engine="python")

        else:

            excel = pd.ExcelFile(rk_file)

            sheet = detect_rk_sheet(excel)

            header_row = detect_header(excel, sheet)

            df = pd.read_excel(excel, sheet_name=sheet, header=header_row)

        df.columns = df.columns.str.strip()

        id_col = detect_id_column(df)

        desc_col = detect_desc_column(df)

        df["KODE_UNIK"] = df[desc_col].apply(ambil_kode_unik)

        database = df[[id_col,"KODE_UNIK",desc_col]].copy()

        database.columns = ["ID","KODE_UNIK","Uraian Transaksi"]

        database["ID"] = pd.to_numeric(database["ID"], errors="coerce")

        database = database.dropna(subset=["ID"])

        database = database[
            ~((database["KODE_UNIK"]=="N/A") & (database["Uraian Transaksi"].isna()))
        ]

        valid = database[database["KODE_UNIK"]!="N/A"].copy()
        na_data = database[database["KODE_UNIK"]=="N/A"].copy()

        valid = valid.drop_duplicates(subset=["ID","KODE_UNIK"])

        new_data = valid

        if existing_db_file:

            excel_db = pd.ExcelFile(existing_db_file)

            sheet_db = excel_db.sheet_names[0]

            old_db = pd.read_excel(excel_db, sheet_name=sheet_db)

            old_db["ID"] = old_db["ID"].astype(str)
            old_db["KODE_UNIK"] = old_db["KODE_UNIK"].astype(str)

            existing_keys = set(zip(old_db["ID"],old_db["KODE_UNIK"]))

            new_data = valid[
                ~valid.apply(
                    lambda r: (str(r["ID"]),str(r["KODE_UNIK"])) in existing_keys,
                    axis=1
                )
            ]

        normal, conflict = group_conflict(new_data)

        if not normal.empty:
            normal = normal.sort_values("ID")

        if not conflict.empty:
            conflict = conflict.sort_values("ID")

        col1,col2,col3,col4 = st.columns(4)

        col1.metric("Total transaksi",len(database))
        col2.metric("Data normal",len(normal))
        col3.metric("Data konflik",len(conflict))
        col4.metric("N/A",len(na_data))

        st.success("Database berhasil dibuat")

        wb = Workbook()
        ws = wb.active

        ws.append(["ID","KODE_UNIK","Uraian Transaksi"])

        yellow = PatternFill(start_color="FFFF00",fill_type="solid")
        red = PatternFill(start_color="FF9999",fill_type="solid")

        for _,row in normal.iterrows():
            ws.append(row.tolist())

        for _,row in conflict.iterrows():

            ws.append(row.tolist())

            for c in ws[ws.max_row]:
                c.fill = yellow

        for _,row in na_data.iterrows():

            ws.append(row.tolist())

            for c in ws[ws.max_row]:
                c.fill = red

        output = BytesIO()
        wb.save(output)

        st.download_button(
            "Download DATABASE_BRI.xlsx",
            output.getvalue(),
            "DATABASE_BRI.xlsx"
        )

    except Exception as e:

        st.error("Terjadi error saat memproses file")
        st.write(e)
